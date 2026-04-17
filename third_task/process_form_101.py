"""
process_form_101.py — расчёт чистой прибыли банков по форме 101 ЦБ РФ

Скрипт готовит данные и кладёт их в Excel-файл с двумя листами:
  • data        — полные данные по всем банкам + именованные диапазоны
                  (`bank_names`, `profit_abs`, `profit_mom`, `profit_yoy`),
                  поверх которых удобно собрать дашборд
  • validation  — сводка сверки расчёта с формой 102 Q1 2021

Сам интерактивный дашборд (выпадающий список, Top-10, диаграмма)
собирается в Excel вручную поверх этих именованных диапазонов.
Пошаговая инструкция — в README.md («Сборка дашборда»)

Что делает:
  1. Читает DBF-файлы формы 101 за 5 отчётных дат (март–май 2021 и апрель–май 2020).
  2. По каждому банку суммирует счета 70601–70605, 70613, 70615 (+)
     и вычитает 70606–70611, 70614, 70616 (−) — колонка IITG.
     Значения отсутствующих счетов не влияют на результат
  3. Из кумулятивных (YTD) значений получает месячные прибыли:
        прибыль_месяца(t) = YTD(t) − YTD(t−1)
  4. Считает прирост прибыли за месяц (май 2021 vs апрель 2021)
     и прирост за год (май 2021 vs май 2020)
  5. Пришивает названия банков из справочника N1.DBF
  6. Сверяет результат с формой 102 Q1 2021 (счета 61101/61102) — валидация метода

Единицы: IITG в форме 101 ЦБ РФ выражен в тысячах рублей.
В итоговом Excel значения представлены в МЛРД рублей (делим на 1 000 000) —
привычный масштаб для банковских прибылей

Скрипт работает полностью офлайн при наличии DBF-файлов в data/raw/.
"""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from dbfread import DBF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName

# ── Пути и константы ────────────────────────────────────────────────────────

BASE_DIR = Path(__file__).resolve().parent
RAW_DIR = BASE_DIR / "data" / "raw"
OUTPUT_XLSX = BASE_DIR / "bank_profit_report.xlsx"

# Счета формы 101, из которых собирается чистая прибыль
# Логика ТЗ: сумма положительных минус сумма отрицательных
POSITIVE_ACCOUNTS: set[str] = {"70601", "70602", "70603", "70604", "70605", "70613", "70615"}
NEGATIVE_ACCOUNTS: set[str] = {"70606", "70607", "70608", "70609", "70610", "70611", "70614", "70616"}

# Периоды и соответствующие DBF-файлы формы 101 (YTD)
# Ключ — код периода для внутреннего использования
FORM_101_FILES: dict[str, Path] = {
    "2020-04": RAW_DIR / "042020B1.DBF",
    "2020-05": RAW_DIR / "052020B1.DBF",
    "2021-03": RAW_DIR / "032021B1.DBF",
    "2021-04": RAW_DIR / "042021B1.DBF",
    "2021-05": RAW_DIR / "052021B1.DBF",
}

# Справочник названий банков. Берём самый свежий из скачанных — он покрывает все
# интересующие нас банки
BANK_NAMES_FILE: Path = RAW_DIR / "052021N1.DBF"

# Форма 102 Q1 2021 для сверки расчёта.
FORM_102_Q1_FILE: Path = RAW_DIR / "12021_P1.DBF"
VALIDATION_POS_CODE = "61101"
VALIDATION_NEG_CODE = "61102"

# Коэффициент для перевода в млрд ₽ (IITG хранится в тыс. ₽).
# тыс. → млрд  ⇒  делим на 1 000 000.
THOUSANDS_TO_BILLIONS = 1_000_000


# ── Разбор DBF формы 101 ────────────────────────────────────────────────────

def compute_ytd_profit(dbf_path: Path) -> dict[int, float]:
    """
    Вернуть словарь REGN → накопленная (YTD) прибыль в тыс. ₽
    на конец периода, соответствующего файлу dbf_path.

    Суммируем IITG положительных счетов, вычитаем IITG отрицательных.
    Если у банка отсутствует часть счетов из списка — ничего страшного,
    складываем то, что есть.
    """
    profit: dict[int, float] = defaultdict(float)
    for row in DBF(str(dbf_path), encoding="cp866", load=False):
        account = str(row["NUM_SC"]).strip()
        if account in POSITIVE_ACCOUNTS:
            profit[row["REGN"]] += float(row["IITG"] or 0)
        elif account in NEGATIVE_ACCOUNTS:
            profit[row["REGN"]] -= float(row["IITG"] or 0)
    return dict(profit)


def load_bank_names(dbf_path: Path) -> dict[int, str]:
    """Справочник REGN -> «АО ЮниКредит Банк» и т. п."""
    names: dict[int, str] = {}
    for row in DBF(str(dbf_path), encoding="cp866", load=False):
        regn = row["REGN"]
        name = (row.get("NAME_B") or "").strip()
        if regn and name:
            names[regn] = name
    return names


# ── Разбор формы 102 для валидации ──────────────────────────────────────────

def compute_form_102_profit(dbf_path: Path) -> dict[int, float]:
    """
    По форме 102 рассчитать прибыль: cумма(код 61101) − сумма(код 61102) по колонке SIM_ITOGO.
    Единицы — тыс. руб. (как и в форме 101)
    """
    profit: dict[int, float] = defaultdict(float)
    for row in DBF(str(dbf_path), encoding="cp866", load=False):
        code = str(row["CODE"]).strip()
        value = float(row["SIM_ITOGO"] or 0)
        if code == VALIDATION_POS_CODE:
            profit[row["REGN"]] += value
        elif code == VALIDATION_NEG_CODE:
            profit[row["REGN"]] -= value
    return dict(profit)


# ── Основная подготовка данных ──────────────────────────────────────────────

def build_bank_dataset() -> tuple[list[dict], dict[str, float]]:
    """
    Собираем строки для листа data и метрики для листа validation

    Функция возвращает:
      rows       — список словарей, по одному на банк,
                   с полями regn, name, profit_may_2021_mln, profit_apr_2021_mln,
                   profit_may_2020_mln, mom_mln, yoy_mln
      validation — сводные метрики сверки с формой 102
    """
    print("Чтение DBF-файлов формы 101…")
    ytd: dict[str, dict[int, float]] = {}
    for period, path in FORM_101_FILES.items():
        print(f"  {period}: {path.name}")
        ytd[period] = compute_ytd_profit(path)

    print(f"Чтение справочника банков: {BANK_NAMES_FILE.name}")
    names = load_bank_names(BANK_NAMES_FILE)

    # Объединение REGN из всех пяти файлов — банк должен учитываться,
    # даже если по какому-то периоду данных нет (получит прочерк в Excel).
    all_regns: set[int] = set()
    for period_data in ytd.values():
        all_regns.update(period_data.keys())

    # Помесячные прибыли: разница соседних YTD.
    # Для мая 2021 база — апрель 2021 (тот же год, YTD прост).
    # Для апреля 2021 база — март 2021.
    # Для мая 2020 база — апрель 2020.
    rows: list[dict] = []
    for regn in sorted(all_regns):
        name = names.get(regn, f"[REGN {regn}]")

        def monthly(period: str, prev_period: str) -> float | None:
            """Месячная прибыль: разница YTD соседних периодов. None, если данных нет"""
            if regn not in ytd[period] or regn not in ytd[prev_period]:
                return None
            return (ytd[period][regn] - ytd[prev_period][regn]) / THOUSANDS_TO_BILLIONS

        may_2021 = monthly("2021-05", "2021-04")
        apr_2021 = monthly("2021-04", "2021-03")
        may_2020 = monthly("2020-05", "2020-04")

        mom = (may_2021 - apr_2021) if (may_2021 is not None and apr_2021 is not None) else None
        yoy = (may_2021 - may_2020) if (may_2021 is not None and may_2020 is not None) else None

        rows.append({
            "regn": regn,
            "name": name,
            "profit_may_2021": may_2021,
            "profit_apr_2021": apr_2021,
            "profit_may_2020": may_2020,
            "mom": mom,
            "yoy": yoy,
        })

    # Валидация: сверка YTD формы 101 на 01.04.2021 с формой 102 Q1 2021
    print(f"Валидация через форму 102: {FORM_102_Q1_FILE.name}")
    form_102_profit = compute_form_102_profit(FORM_102_Q1_FILE)
    form_101_ytd_q1 = ytd["2021-03"]

    matches = 0
    minor_diff = 0  # расхождение < 1 %
    major_diff = 0  # расхождение ≥ 1 %
    missing = 0
    max_abs_diff = 0.0
    max_abs_diff_regn: int | None = None

    for regn in form_101_ytd_q1:
        v101 = form_101_ytd_q1[regn]
        v102 = form_102_profit.get(regn)
        if v102 is None:
            missing += 1
            continue
        diff = abs(v101 - v102)
        denom = max(abs(v101), abs(v102), 1.0)
        if diff / denom < 1e-6:
            matches += 1
        elif diff / denom < 0.01:
            minor_diff += 1
        else:
            major_diff += 1
        if diff > max_abs_diff:
            max_abs_diff = diff
            max_abs_diff_regn = regn

    validation = {
        "total_banks_form_101_q1": len(form_101_ytd_q1),
        "exact_matches": matches,
        "minor_diff_<1pct": minor_diff,
        "major_diff_>=1pct": major_diff,
        "not_found_in_form_102": missing,
        "max_abs_diff_thousands_rub": max_abs_diff,
        "max_abs_diff_regn": max_abs_diff_regn,
    }
    return rows, validation


# ── Сборка Excel ────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def write_data_sheet(wb: Workbook, rows: list[dict]) -> int:
    """Лист data с полными данными по всем банкам. Возвращает номер последней строки"""
    ws = wb.create_sheet("data")
    headers = [
        "REGN",
        "Банк",
        "Абс. прибыль, май 2021, млрд ₽",
        "Прибыль, апрель 2021, млрд ₽",
        "Прибыль, май 2020, млрд ₽",
        "MoM прирост, млрд ₽",
        "YoY прирост, млрд ₽",
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in rows:
        ws.append([
            r["regn"],
            r["name"],
            r["profit_may_2021"],
            r["profit_apr_2021"],
            r["profit_may_2020"],
            r["mom"],
            r["yoy"],
        ])

    last_row = ws.max_row

    # Ширина колонок и числовой формат. 3 знака после запятой, чтобы в млрд 
    # читаемо отображались и мелкие банки (0,001 млрд = 1 млн рублей)
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 55
    for col_letter in ("C", "D", "E", "F", "G"):
        ws.column_dimensions[col_letter].width = 22
        for row_idx in range(2, last_row + 1):
            ws[f"{col_letter}{row_idx}"].number_format = "#,##0.000;-#,##0.000;-"

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:G{last_row}"

    # Именованные диапазоны для удобных формул на дашборде
    last = last_row
    defined_names = {
        "bank_names": f"data!$B$2:$B${last}",
        "profit_abs": f"data!$C$2:$C${last}",
        "profit_mom": f"data!$F$2:$F${last}",
        "profit_yoy": f"data!$G$2:$G${last}",
    }
    for name, ref in defined_names.items():
        wb.defined_names[name] = DefinedName(name, attr_text=ref)

    return last_row


def write_validation_sheet(wb: Workbook, validation: dict) -> None:
    """Лист validation — итоги сверки формы 101 с формой 102 Q1 2021."""
    ws = wb.create_sheet("validation")
    ws["A1"] = "Сверка расчёта прибыли: форма 101 vs форма 102, Q1 2021"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E78")

    ws["A3"] = "Метод"
    ws["B3"] = (
        "Для каждого банка: YTD-прибыль по форме 101 на 01.04.2021 "
        "(счета 70601–70605, 70613, 70615 минус 70606–70611, 70614, 70616) "
        "сравнивается с прибылью по форме 102 Q1 2021 (коды 61101 минус 61102)."
    )
    ws["B3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[3].height = 45
    ws.merge_cells("B3:D3")

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 20

    ws["A5"] = "Всего банков (форма 101, Q1)"
    ws["B5"] = validation["total_banks_form_101_q1"]

    ws["A6"] = "Точные совпадения"
    ws["B6"] = validation["exact_matches"]

    ws["A7"] = "Расхождение < 1 %"
    ws["B7"] = validation["minor_diff_<1pct"]

    ws["A8"] = "Расхождение ≥ 1 %"
    ws["B8"] = validation["major_diff_>=1pct"]

    ws["A9"] = "Нет в форме 102"
    ws["B9"] = validation["not_found_in_form_102"]

    ws["A11"] = "Макс. абс. расхождение, тыс. ₽"
    ws["B11"] = validation["max_abs_diff_thousands_rub"]
    ws["B11"].number_format = "#,##0.00"
    ws["A12"] = "REGN банка с макс. расхождением"
    ws["B12"] = validation["max_abs_diff_regn"]

    for r in range(5, 13):
        ws[f"A{r}"].font = Font(bold=True)


def build_excel(rows: list[dict], validation: dict) -> None:
    """Собрать итоговый .xlsx: data + validation. Дашборд — вручную в Excel"""
    wb = Workbook()
    # openpyxl создаёт дефолтный лист; удаляем его, заменим на наши
    default = wb.active
    wb.remove(default)

    write_data_sheet(wb, rows)
    write_validation_sheet(wb, validation)

    # Флаг пересчёта оставляем на случай, если при ручной сборке дашборда
    # сами добавим формулы, Excel тогда сам пересчитает при открытии
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True

    wb.save(OUTPUT_XLSX)
    print(f"Итоговый файл: {OUTPUT_XLSX}")
    print(
        "Дашборд собирается в Excel вручную поверх именованных диапазонов "
        "`bank_names`, `profit_abs`, `profit_mom`, `profit_yoy` "
        "на листе `data`. Инструкция — в README.md."
    )



def main() -> int:
    missing = [p for p in list(FORM_101_FILES.values()) + [BANK_NAMES_FILE, FORM_102_Q1_FILE] if not p.exists()]
    if missing:
        print("ОШИБКА: не хватает файлов, запустите сначала download_data.py")
        for p in missing:
            print(f"  нет: {p}")
        return 1

    rows, validation = build_bank_dataset()

    # Выводим краткую сводку в CLI 
    banks_with_may = [r for r in rows if r["profit_may_2021"] is not None]
    print(f"\nВсего банков с данными за май 2021: {len(banks_with_may)}")
    top5 = sorted(banks_with_may, key=lambda r: r["profit_may_2021"], reverse=True)[:5]
    print("Top-5 по абсолютной прибыли за май 2021 (млрд ₽):")
    for r in top5:
        print(f"  {r['regn']:>5}  {r['name'][:50]:<50}  {r['profit_may_2021']:>15,.0f}")

    print("\nВалидация (форма 101 vs форма 102, Q1 2021):")
    for k, v in validation.items():
        print(f"  {k}: {v}")

    build_excel(rows, validation)
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
